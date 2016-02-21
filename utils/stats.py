from openpyxl import Workbook
from subprocess import call
import common

columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
wb = Workbook()
ws = wb.active

def write_header(ws, row):
	headers = ['Threads', 'Run 1', 'Run 2', 'Run 3', 'Run 4', 'Run 5', 'Average']

	# Setup Headers
	for i in range(7):
		ws[columns[i]+str(row)] = headers[i]

def run(ws, starting_row, problem_size):
	write_header(ws, starting_row)
	common.run_datagen(problem_size)
	counter = 0;

	for i in range(1, min(problem_size+1, 201)):
		if(problem_size % i == 0):
			counter += 1
			average = 0
			ws['A'+str(starting_row + counter)] = i
			
			for j in range(5):
				common.run_main(i)
				runtime = common.get_runtime()
				ws[columns[j+1]+str(starting_row + counter)] = runtime
				average += runtime

			average = average / 5
			ws[columns[6]+str(starting_row + counter)] = average

	
	common.cleanup()
	return counter + 2;

def multi_run(sizes):
	start = 1
	for i in sizes:
		start += run(ws, start, i)

multi_run([500, 1000])
common.cleanup()
wb.save("stats.xlsx")